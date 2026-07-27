#!/usr/bin/env python3
"""
mspaudit.py — findings engine for a school-finance workbook estate.

WHY THIS EXISTS.

A state minimum-school-program calculation is not a spreadsheet. It is a
distributed calculation graph across dozens of workbooks, wired by external
links, carrying statutory coefficients to five decimals, with no version
control, no dependency map, and no test harness. The authority for each
constant lives in code and rule; the constant itself lives in a cell; and the
mapping between them lives in whichever person happens to know.

That arrangement fails in one direction. It does not produce small errors. It
produces a nine-figure misstatement that nothing detects until a district
telephones to report money it did not expect — a control environment whose sole
detective control is a recipient acting against its own interest.

This tool is the check that should have existed. It reads the estate, builds the
graph, and reports the defects that precede that failure.

    python mspaudit.py <folder>                 audit an estate
    python mspaudit.py <folder> --json out.json machine-readable findings
    python mspaudit.py <folder> --md report.md  a report you can hand a board

WHAT IT FINDS

  F1  FORMULA INCONSISTENCY IN A BLOCK — a column of 200 identical formulas
      with one that differs. This is how the money moves wrong. It is the
      single highest-yield finding in spreadsheet audit and the hardest to
      see by eye.
  F2  UNRESOLVED EXTERNAL LINK — a formula reaching into a workbook that is
      not in the estate. The value it returns is whatever was cached the last
      time somebody opened it, which may be a prior fiscal year.
  F3  STALE-YEAR REFERENCE — an FY22 workbook wired to a file named FY20.
      Sometimes correct, always worth a written reason.
  F4  UNDOCUMENTED COEFFICIENT — a statutory-looking constant sitting in a
      formula with no annotation naming the section it implements.
  F5  VERSION-AMBIGUOUS INPUT — several files differing only by a date or
      version suffix, where nothing records which one fed the distribution.
  F6  FAN-IN HOTSPOT — a cell many others depend on. Not a defect; a single
      point of failure that should be named, reconciled, and owned.
  F7  HARD-CODED OVERRIDE — a formula whose result is partly a literal,
      inside a sheet that otherwise looks up its values.

Nothing here is proprietary to one state. Every state running a weighted-pupil
formula has this estate and this exposure.

Standard library plus openpyxl.
"""

import argparse
import json
import re
import sys
from collections import defaultdict, Counter
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("FAIL — needs openpyxl:  pip install openpyxl"); sys.exit(1)

CELL = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d{1,7})")
EXTIDX = re.compile(r"\[(\d+)\]")
NUMLIT = re.compile(r"(?<![A-Z0-9_.$!:])(\d+\.\d{3,}|0\.\d{2,})(?![0-9])")
YEAR = re.compile(r"(?:FY|SY)\s?(\d{2,4})", re.I)
VERSUFFIX = re.compile(r"[ _\-.](\d{1,2}[._-]\d{1,2}[._-]\d{2,4}|v\d+|"
                       r"final|draft|old|copy|rev\d*|\(\d+\))", re.I)


def col_to_n(c):
    n = 0
    for ch in c:
        n = n * 26 + (ord(ch) - 64)
    return n


def normalise(formula, row, col):
    """Rewrite a formula into offset form so two cells doing 'the same thing'
    compare equal. =SUM(C3:C43) in C44 and =SUM(D3:D43) in D44 are the same
    formula; =SUM(C3:C42) is not, and that difference is the finding."""
    def sub(m):
        adol, cl, rdol, rw = m.groups()
        c2 = col_to_n(cl) if adol else col_to_n(cl) - col
        r2 = int(rw) if rdol else int(rw) - row
        return f"R[{r2}]C[{c2}]"
    return CELL.sub(sub, formula)


def scan(folder: Path):
    books = {}
    unreadable = []
    files = [p for p in sorted(folder.rglob("*"))
             if p.suffix.lower() in (".xlsx", ".xlsm") and not p.name.startswith("~$")]
    legacy = [p for p in sorted(folder.rglob("*")) if p.suffix.lower() == ".xls"]
    for p in files:
        try:
            wb = openpyxl.load_workbook(p, data_only=False, keep_links=True)
        except Exception as e:
            unreadable.append((p, str(e)[:70])); continue
        ext = []
        for el in (getattr(wb, "_external_links", None) or []):
            try:
                ext.append(el.file_link.Target)
            except Exception:
                ext.append("?")
        sheets = {}
        for ws in wb.worksheets:
            cells = {}
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and c.value.startswith("="):
                        cells[(c.row, c.column)] = c.value
            if cells:
                sheets[ws.title] = cells
        books[p] = {"sheets": sheets, "ext": ext}
    return books, legacy, unreadable


# ----------------------------------------------------------------------
def find_inconsistency(books, minrun=6, tol=0.80):
    out = []
    for p, b in books.items():
        for sh, cells in b["sheets"].items():
            bycol = defaultdict(list)
            for (r, c), f in cells.items():
                bycol[c].append((r, f))
            for c, rows in bycol.items():
                rows.sort()
                run = []
                for r, f in rows + [(None, None)]:
                    if run and r is not None and r == run[-1][0] + 1:
                        run.append((r, f)); continue
                    if len(run) >= minrun:
                        pats = [(r_, normalise(f_, r_, c)) for r_, f_ in run]
                        cnt = Counter(x[1] for x in pats)
                        top, n = cnt.most_common(1)[0]
                        if n / len(pats) >= tol and n < len(pats):
                            odd = [r_ for r_, pt in pats if pt != top]
                            # A subtotal at the head or foot of a block is SUPPOSED
                            # to differ from the rows it totals. Flagging those is
                            # the classic false positive that gets an audit ignored
                            # by page three. Only interior deviations are findings:
                            # a cell surrounded on both sides by the majority
                            # formula and not matching it has no structural excuse.
                            edges = {run[0][0], run[-1][0]}
                            odd = [r_ for r_ in odd if r_ not in edges]
                            if not odd:
                                run = [(r, f)] if r is not None else []
                                continue
                            out.append({"finding": "F1", "severity": "high",
                                        "file": p.name, "sheet": sh,
                                        "column": openpyxl.utils.get_column_letter(c),
                                        "block_rows": [run[0][0], run[-1][0]],
                                        "majority": n, "of": len(pats),
                                        "odd_rows": odd[:12],
                                        "example": dict(run)[odd[0]][:160]})
                    run = [(r, f)] if r is not None else []
    return out


def find_links(books):
    names = {p.name.lower() for p in books}
    out = []
    for p, b in books.items():
        used = set()
        for sh, cells in b["sheets"].items():
            for f in cells.values():
                for i in EXTIDX.findall(f):
                    used.add(int(i))
        for i in sorted(used):
            tgt = b["ext"][i - 1] if 0 < i <= len(b["ext"]) else None
            base = Path(tgt).name if tgt else None
            if not base:
                out.append({"finding": "F2", "severity": "high", "file": p.name,
                            "link_index": i, "target": None,
                            "note": "external link index used in formulas but no link "
                                    "record resolves it"})
            elif base.lower() not in names:
                out.append({"finding": "F2", "severity": "high", "file": p.name,
                            "link_index": i, "target": base,
                            "note": "target workbook is not in the estate; the value "
                                    "used is whatever Excel cached"})
            else:
                sy = YEAR.findall(p.name); ty = YEAR.findall(base)
                if sy and ty and sy[0][-2:] != ty[0][-2:]:
                    out.append({"finding": "F3", "severity": "medium", "file": p.name,
                                "target": base,
                                "note": f"FY/SY {sy[0]} workbook is wired to an "
                                        f"FY/SY {ty[0]} source — often correct, always "
                                        f"worth a written reason"})
    return out


def find_coefficients(books):
    out = []
    for p, b in books.items():
        for sh, cells in b["sheets"].items():
            hits = Counter()
            example = {}
            for (r, c), f in cells.items():
                for lit in NUMLIT.findall(f):
                    hits[lit] += 1
                    example.setdefault(lit, (r, c, f))
            for lit, n in hits.most_common(6):
                if n >= 3:
                    r, c, f = example[lit]
                    out.append({"finding": "F4", "severity": "medium", "file": p.name,
                                "sheet": sh, "constant": lit, "occurrences": n,
                                "example_cell": f"{openpyxl.utils.get_column_letter(c)}{r}",
                                "example": f[:150],
                                "note": "high-precision constant repeated in formulas. If "
                                        "this is statutory it should name its authority "
                                        "and live in one cell, not many."})
    return out


def find_versions(folder):
    groups = defaultdict(list)
    for p in sorted(folder.rglob("*")):
        if p.suffix.lower() not in (".xls", ".xlsx", ".xlsm"):
            continue
        if p.name.startswith("~$"):
            continue
        key = VERSUFFIX.sub("", p.stem).strip().lower()
        groups[key].append(p.name)
    return [{"finding": "F5", "severity": "medium", "stem": k, "files": v,
             "note": "several files differ only by a date or version suffix; nothing "
                     "in the estate records which one fed the distribution"}
            for k, v in groups.items() if len(v) > 1]


def find_hotspots(books, top=8):
    fan = Counter()
    for p, b in books.items():
        for sh, cells in b["sheets"].items():
            for f in cells.values():
                for m in CELL.finditer(f):
                    fan[(p.name, sh, m.group(2) + m.group(4))] += 1
    return [{"finding": "F6", "severity": "info", "file": k[0], "sheet": k[1],
             "cell": k[2], "referenced_by": n,
             "note": "single point of failure: many cells depend on this one. Name an "
                     "owner and a reconciliation, or it is an undocumented control."}
            for k, n in fan.most_common(top) if n >= 25]


def find_hardcoded(books):
    out = []
    for p, b in books.items():
        for sh, cells in b["sheets"].items():
            lookups = sum(1 for f in cells.values()
                          if "VLOOKUP" in f or "INDEX(" in f or "GETPIVOTDATA" in f)
            if lookups < 5:
                continue
            for (r, c), f in cells.items():
                if re.search(r"[+\-*/]\s*\d{4,}(?![0-9.])", f) and "VLOOKUP" not in f:
                    out.append({"finding": "F7", "severity": "medium", "file": p.name,
                                "sheet": sh,
                                "cell": f"{openpyxl.utils.get_column_letter(c)}{r}",
                                "example": f[:150],
                                "note": "literal arithmetic on a sheet that otherwise "
                                        "looks its values up — an override with no "
                                        "recorded reason"})
                    break
    return out


# ----------------------------------------------------------------------
SEV = {"high": 0, "medium": 1, "info": 2}
TITLES = {"F1": "Formula inconsistency inside a block",
          "F2": "Unresolved external link",
          "F3": "Stale-year reference",
          "F4": "Undocumented coefficient",
          "F5": "Version-ambiguous input",
          "F6": "Fan-in hotspot",
          "F7": "Hard-coded override"}


def main():
    ap = argparse.ArgumentParser(description="mspaudit — school-finance workbook audit")
    ap.add_argument("folder")
    ap.add_argument("--json"); ap.add_argument("--md")
    ap.add_argument("--quiet", action="store_true")
    a = ap.parse_args()
    folder = Path(a.folder)
    if not folder.exists():
        print(f"FAIL — no such folder: {folder}"); return 1

    books, legacy, unreadable = scan(folder)
    nf = sum(len(c) for b in books.values() for c in b["sheets"].values())
    ns = sum(len(b["sheets"]) for b in books.values())

    findings = (find_inconsistency(books) + find_links(books) +
                find_coefficients(books) + find_versions(folder) +
                find_hotspots(books) + find_hardcoded(books))
    findings.sort(key=lambda f: (SEV[f["severity"]], f["finding"]))

    if not a.quiet:
        print(f"\nMSP WORKBOOK AUDIT — {folder}")
        print("=" * 70)
        print(f"  {len(books)} workbook(s) parsed · {ns} sheet(s) · {nf:,} formulas")
        if legacy:
            print(f"  {len(legacy)} legacy .xls not parsed — convert to .xlsx to include "
                  f"them; a format nothing can read is itself a finding")
        for p, e in unreadable:
            print(f"  UNREADABLE {p.name}: {e}")
        c = Counter(f["finding"] for f in findings)
        print("\n  FINDINGS")
        for k in ("F1", "F2", "F3", "F4", "F5", "F6", "F7"):
            if c[k]:
                print(f"    {k}  {c[k]:>4}  {TITLES[k]}")
        print()
        for f in findings[:24]:
            head = f"  [{f['finding']}·{f['severity']}] {f.get('file','')}"
            if f.get("sheet"):
                head += f" :: {f['sheet']}"
            print(head)
            if f["finding"] == "F1":
                print(f"      column {f['column']} rows {f['block_rows'][0]}-"
                      f"{f['block_rows'][1]}: {f['majority']} of {f['of']} share one "
                      f"formula. Odd: {f['odd_rows']}")
                print(f"      {f['example']}")
            else:
                for k in ("target", "constant", "occurrences", "cell", "example_cell",
                          "referenced_by", "stem", "files", "example"):
                    if f.get(k) is not None and k in f:
                        print(f"      {k}: {f[k]}")
                print(f"      {f['note']}")
        if len(findings) > 24:
            print(f"\n  … {len(findings)-24} more. Use --json or --md for the full set.")
        print()

    if a.json:
        Path(a.json).write_text(json.dumps({
            "tool": "mspaudit/1.0",
            "as_of": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "estate": str(folder), "workbooks": len(books), "sheets": ns,
            "formulas": nf, "legacy_xls_not_parsed": [p.name for p in legacy],
            "findings": findings}, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8")
        print(f"  json → {a.json}")

    if a.md:
        L = [f"# Workbook audit — {folder.name}", "",
             f"*{datetime.now(timezone.utc):%Y-%m-%d} · mspaudit/1.0*", "",
             f"{len(books)} workbooks, {ns} sheets, {nf:,} formulas.", "",
             "This audit reports structure, not opinion. Every finding names a file, a "
             "sheet and a cell, and can be checked without trusting the auditor.", ""]
        c = Counter(f["finding"] for f in findings)
        L += ["## Summary", "", "| code | count | finding |", "|---|---|---|"]
        for k in ("F1", "F2", "F3", "F4", "F5", "F6", "F7"):
            if c[k]:
                L.append(f"| {k} | {c[k]} | {TITLES[k]} |")
        for k in ("F1", "F2", "F3", "F4", "F5", "F6", "F7"):
            rows = [f for f in findings if f["finding"] == k]
            if not rows:
                continue
            L += ["", f"## {k} — {TITLES[k]}", ""]
            for f in rows[:40]:
                loc = f.get("file", "")
                if f.get("sheet"):
                    loc += f" :: {f['sheet']}"
                if k == "F1":
                    L.append(f"- **{loc}** column {f['column']}, rows "
                             f"{f['block_rows'][0]}–{f['block_rows'][1]} — "
                             f"{f['majority']} of {f['of']} identical, odd rows "
                             f"{f['odd_rows']}. `{f['example']}`")
                else:
                    L.append(f"- **{loc}** — " + f.get("note", ""))
            if len(rows) > 40:
                L.append(f"- … {len(rows)-40} more")
        Path(a.md).write_text("\n".join(L) + "\n", encoding="utf-8")
        print(f"  markdown → {a.md}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
